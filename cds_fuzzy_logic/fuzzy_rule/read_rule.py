import openpyxl

def read_throttle_rule(): 
    throttle_rule = []
    book = openpyxl.load_workbook(r'cds_fuzzy_logic/fuzzy_rule/rule(main).xlsx')
    sheet = book.worksheets[0]

    steering = [cell[0] for cell in sheet.iter_rows(min_col=2, min_row=1, max_row=4, values_only=True)]
    speed = [cell[0] for cell in sheet.iter_rows(min_col=3, min_row=1, max_row=4, values_only=True)]

    for i in range(1, len(steering)):
        throttle_rule.append((steering[i], speed[i]))

    return throttle_rule

def read_impediment_rule():
    impediment_rule = []
    book = openpyxl.load_workbook(r'cds_fuzzy_logic/fuzzy_rule/rule(main).xlsx')
    sheet = book.worksheets[1]

    distance = [cell[0] for cell in sheet.iter_rows(min_col=2, min_row=1, max_row=10, values_only=True)]
    steering = [cell[0] for cell in sheet.iter_rows(min_col=3, min_row=1, max_row=10, values_only=True)]
    speed = [cell[0] for cell in sheet.iter_rows(min_col=4, min_row=1, max_row=10, values_only=True)]

    for i in range(1, len(distance)):
        impediment_rule.append((distance[i], steering[i], speed[i]))

    return impediment_rule

def read_straight_rule():
    straight_rule = []
    book = openpyxl.load_workbook(r'cds_fuzzy_logic/fuzzy_rule/rule(main).xlsx')
    sheet = book.worksheets[2]

    distance = [cell[0] for cell in sheet.iter_rows(min_col=2, min_row=1, max_row=10, values_only=True)]
    steering = [cell[0] for cell in sheet.iter_rows(min_col=3, min_row=1, max_row=10, values_only=True)]
    speed = [cell[0] for cell in sheet.iter_rows(min_col=4, min_row=1, max_row=10, values_only=True)]

    for i in range(1, len(distance)):
        straight_rule.append((distance[i], steering[i], speed[i]))

    return straight_rule

def read_stop_rule():
    stop_rule = []
    book = openpyxl.load_workbook(r'cds_fuzzy_logic/fuzzy_rule/rule(main).xlsx')
    sheet = book.worksheets[3]

    distance = [cell[0] for cell in sheet.iter_rows(min_col=2, min_row=1, max_row=10, values_only=True)]
    steering = [cell[0] for cell in sheet.iter_rows(min_col=3, min_row=1, max_row=10, values_only=True)]
    speed = [cell[0] for cell in sheet.iter_rows(min_col=4, min_row=1, max_row=10, values_only=True)]

    for i in range(1, len(distance)):
        stop_rule.append((distance[i], steering[i], speed[i]))

    return stop_rule

def read_lr_rule():
    lr_rule = []
    book = openpyxl.load_workbook(r'cds_fuzzy_logic/fuzzy_rule/rule(main).xlsx')
    sheet = book.worksheets[4]

    distance = [cell[0] for cell in sheet.iter_rows(min_col=2, min_row=1, max_row=10, values_only=True)]
    steering = [cell[0] for cell in sheet.iter_rows(min_col=3, min_row=1, max_row=10, values_only=True)]
    speed = [cell[0] for cell in sheet.iter_rows(min_col=4, min_row=1, max_row=10, values_only=True)]

    for i in range(1, len(distance)):
        lr_rule.append((distance[i], steering[i], speed[i]))

    return lr_rule

def read_noentry_rule():
    noentry_rule = []
    book = openpyxl.load_workbook(r'cds_fuzzy_logic/fuzzy_rule/rule(main).xlsx')
    sheet = book.worksheets[5]

    distance = [cell[0] for cell in sheet.iter_rows(min_col=2, min_row=1, max_row=10, values_only=True)]
    steering = [cell[0] for cell in sheet.iter_rows(min_col=3, min_row=1, max_row=10, values_only=True)]
    speed = [cell[0] for cell in sheet.iter_rows(min_col=4, min_row=1, max_row=10, values_only=True)]

    for i in range(1, len(distance)):
        noentry_rule.append((distance[i], steering[i], speed[i]))

    return noentry_rule

def read_steering_rule():
    steering_rule = []
    book = openpyxl.load_workbook(r'cds_fuzzy_logic/fuzzy_rule/rule(main).xlsx')
    sheet = book.worksheets[6]

    angle = [cell[0] for cell in sheet.iter_rows(min_col=2, min_row=1, max_row=6, values_only=True)]
    steering = [cell[0] for cell in sheet.iter_rows(min_col=3, min_row=1, max_row=6, values_only=True)]

    for i in range(1, len(angle)):
        steering_rule.append((angle[i], steering[i]))

    return steering_rule

